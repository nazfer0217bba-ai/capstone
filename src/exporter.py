import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_dataframe(test_cases):
    """Converts test cases list to a clean Pandas DataFrame."""
    if not test_cases:
        return pd.DataFrame()
    
    df = pd.DataFrame(test_cases)
    # Rename columns for presentation
    df.columns = [col.replace('_', ' ').title() for col in df.columns]
    # Reorder columns logically
    cols_order = ['Id', 'Element', 'Type', 'Name', 'Description', 'Steps', 'Expected Result', 'Priority']
    cols_order = [c for c in cols_order if c in df.columns]
    return df[cols_order]

def export_to_csv(test_cases):
    """Exports test cases to a CSV string."""
    df = convert_to_dataframe(test_cases)
    if df.empty:
        return ""
    return df.to_csv(index=False)

def export_to_json(test_cases):
    """Exports test cases to a pretty JSON string."""
    if not test_cases:
        return "[]"
    return json.dumps(test_cases, indent=4)

def export_to_excel_writer(test_cases, filepath):
    """
    Exports test cases to an Excel file with premium formatting.
    
    Args:
        test_cases (list of dict): List of test cases.
        filepath (str): Absolute file path to write to.
    """
    df = convert_to_dataframe(test_cases)
    if df.empty:
        return
        
    # Write Excel using Pandas and Openpyxl
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Generated Test Suite')
        
        # Style Sheet
        workbook = writer.book
        worksheet = writer.sheets['Generated Test Suite']
        
        # Define Styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Classic Navy
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        
        thin_side = Side(border_style="thin", color="D3D3D3")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Priority Colors Mapping (optional fill highlighting)
        prio_colors = {
            "CRITICAL": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), # Light Red/Orange
            "HIGH": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),     # Light Yellow
            "MEDIUM": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),   # Light Green
            "LOW": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")       # Light Gray
        }
        
        # Alignments
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
        
        # Apply header styling
        worksheet.row_dimensions[1].height = 26
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = cell_border
            
        # Apply body cell styling
        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 45 # Multi-line wrap space
            priority_val = str(worksheet.cell(row=row_idx, column=df.columns.get_loc('Priority') + 1).value).upper()
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = cell_border
                
                # Apply alignment based on column type
                col_name = df.columns[col_idx - 1]
                if col_name in ['Id', 'Type', 'Priority']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
                # Highlight priority column
                if col_name == 'Priority' and priority_val in prio_colors:
                    cell.fill = prio_colors[priority_val]
                    cell.font = Font(name="Segoe UI", size=10, bold=True)
                    
        # Adjust Column Widths Dynamically
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = col[0].value
            
            # Find the max string length in column
            for cell in col:
                val = str(cell.value or '')
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            
            # Set specific padding based on typical data length
            if col_name in ['Steps', 'Description', 'Expected Result']:
                worksheet.column_dimensions[col_letter].width = 40
            elif col_name in ['Id', 'Priority', 'Type']:
                worksheet.column_dimensions[col_letter].width = 12
            else:
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 15)
                
        # Enable grid lines visibility
        worksheet.views.sheetView[0].showGridLines = True
